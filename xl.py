
import openpyxl
import datetime

class xl:

 #load a workbook from file
 def load_wb(file_name): #file_name (str)
  return openpyxl.load_workbook(file_name)

 #load a worksheet from a worksheet
 def load_ws(work_book, ws_name): #ws_name (str)
  return work_book[ws_name]

 #save a workbook to file
 def save_wb(work_book, file_name):
  work_book.save(file_name)

 #return a cell CONTENT #cells_addr str
 def cell(work_sheet, cell_addr):
  return work_sheet[cell_addr].value

 #assign data to a cell of a worksheet
 def assign(work_sheet, cell_addr, cell_content):
  work_sheet[cell_addr] = cell_content

 #return CONTENT of a column and remove None cell #cot doc
 def column(work_sheet, column_addr):
  content_list = []
  for cell_name in work_sheet[column_addr]:
   if cell_name.value != None:
    content_list = content_list + [str(cell_name.value)]
  return content_list

 #return CONTENT of a column and remove None cell #cot doc
 def column_non_remove(work_sheet, column_addr):
  content_list = []
  for cell_name in work_sheet[column_addr]:
   content_list = content_list + [str(cell_name.value)]
  return content_list


 #return CONTENT of a column-range #row1,2 either string or int
 def col_range(work_sheet, column_addr, row1, row2):
  content = []
  for i in range(int(row1), int(row2)+1):
   if work_sheet[column_addr][i].value != None:
    content = content + [work_sheet[column_addr][i].value]
  return content


 #return corresponding content horizontally
 def cor_content(work_sheet,column1, column2, cell1_value):
  row = xl.row(work_sheet, column1, cell1_value)
  content = work_sheet[column2+str(row)].value
  return content

 #return corresponding content vertically
 def cor_content_col(work_sheet,row1, row2, cell1_value):
  content = []
  for i in range(len(work_sheet[str(row1)])):
   if cell1_value == str(work_sheet[str(row1)][i].value):
    cell_col = chr(i + 65) #65 corresponds to A
  return xl.cell(work_sheet,cell_col + str(row2))


 #return stall names with dishes name
 def stall(work_sheet, column1, column2, cell_value):
  for i in range(len(work_sheet[column1])):
   if cell_value == str(work_sheet[column1][i].value):
    row = i +1
  while str(type(work_sheet[column2][row-1].value)) == "<class 'NoneType'>":
   row = row -1
  return xl.cell(work_sheet, column2 + str(row))

 #return list of sheets
 def sheets(work_book):
  return work_book.get_sheet_names()


 #return the row of a content in a column
 def row(work_sheet, column_addr, content):
  row = 0
  for cell_content in work_sheet[column_addr]:
   if cell_content.value == content:
    row = 1 + work_sheet[column_addr].index(cell_content)
  return row

 #return the row of the next content in a column 
 def next_row(work_sheet, column_addr, content):
  next_row = 0
  next_content = ''
  next_content = xl.column(work_sheet, column_addr)[1 + xl.column(work_sheet, column_addr).index(content)]
  next_row = xl.row(work_sheet, column_addr, next_content)
  return next_row

 #return content of columns in all sheets and remove duplicates
 def all_columns(work_book, column_addr):
  content_list = []
  for work_sheet_name in xl.sheets(work_book):
   work_sheet = work_book[work_sheet_name]
   content_list = content_list + xl.column(work_sheet, column_addr)
  return list(set(content_list))

 #return stall name and sheet name of a food type
 def stall_and_sheet(work_book, column1, column2, cell1_value):
  content_list = []
  for work_sheet_name in xl.sheets(work_book):
   work_sheet = work_book[work_sheet_name]
   for i in range(len(work_sheet[column1])):
    if work_sheet[column1 + str(i+1)].value == cell1_value:
     stall_name = work_sheet[column2 + str(i+1)].value
     content_list = content_list + [str(stall_name) + ' in ' + work_sheet_name]
  return sorted(content_list)

#SAMPLE
print('Excel has loaded!')

