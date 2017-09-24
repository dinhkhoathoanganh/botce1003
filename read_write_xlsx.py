from openpyxl import load_workbook
#load wb from a file
wb = load_workbook('canteen_restaurant_list.xlsx')
#print sheets name
print(wb.get_sheet_names())
# grab the active worksheet
ws = wb.active
#print cells value
print(ws['A1'].value)
# Data can be assigned directly to cells
ws['A1'] = 42
# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()
# Save the file
wb.save("canteen_restaurant_list.xls")
